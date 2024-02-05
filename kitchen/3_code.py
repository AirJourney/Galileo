import json  # 导入json库
import openpyxl


def excel_to_json(excel_file, json_file):
    wb = openpyxl.load_workbook(excel_file)  # 读取excel文件
    excel_data = {}  # 定义字典excel_data存储每个表的数据{表名:数据}
    for sheet in wb.sheetnames:
        result = []  # 定义列表result存储所有读取数据
        for rows in wb[sheet]:  # 获取表的每一行数据
            tmp = {
                'cityCode': '',
                'cityName': ''
            }  # 定义列表tmp存储每一行的数据
            for ci, cell in enumerate(rows):  # 遍历一行每个单元格的数据
                if(ci == 0):
                    tmp['cityCode'] = cell.value
                elif(ci == 1):
                    tmp['cityName'] = cell.value
            result.append(tmp)
        excel_data[sheet] = result
    # 覆盖写入json文件
    with open(json_file, mode='w', encoding='utf-8') as jf:
        json.dump(excel_data, jf, indent=2, sort_keys=True, ensure_ascii=False)


def excel_to_json_by_sheet(excel_file, json_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file)  # 读取excel文件
    excel_data = {}  # 定义字典excel_data存储每个表的数据{表名:数据}
    result = []
    for rows in wb[sheet_name]:  # 获取表的每一行数据
        tmp = {
            'cityCode': '',
            'cityName': ''
        }  # 定义列表tmp存储每一行的数据
        for ci, cell in enumerate(rows):  # 遍历一行每个单元格的数据
            if(ci == 0):
                tmp['cityCode'] = cell.value
            elif(ci == 1):
                tmp['cityName'] = cell.value
        result.append(tmp)
    # 覆盖写入json文件
    with open(json_file, mode='w', encoding='utf-8') as jf:
        json.dump(result, jf, indent=2, sort_keys=True, ensure_ascii=False)


# excel_to_json(r'D:\Github\LLTrip-Cruise\kitchen\top.xlsx',
#               r'D:\Github\LLTrip-Cruise\kitchen\top.json')  # 调用函数，传入参数

# excel_to_json_by_sheet(r'D:\Github\LLTrip-Cruise\kitchen\3code.xlsx',
#               r'D:\Github\LLTrip-Cruise\kitchen\city.json','city')  # 调用函数，传入参数


def readCity1(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file)  # 读取excel文件
    excel_data = {}  # 定义字典excel_data存储每个表的数据{表名:数据}
    result = []
    for rows in wb[sheet_name]:  # 获取表的每一行数据
        tmp = {
            'cityCode': '',
            'cityName': ''
        }  # 定义列表tmp存储每一行的数据
        for ci, cell in enumerate(rows):  # 遍历一行每个单元格的数据
            if(ci == 0):
                tmp['cityCode'] = cell.value
            elif(ci == 1):
                tmp['cityName'] = cell.value
        result.append(tmp)
    return result


def readCity2(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file)  # 读取excel文件
    excel_data = {}  # 定义字典excel_data存储每个表的数据{表名:数据}
    result = []
    for rows in wb[sheet_name]:  # 获取表的每一行数据
        tmp = {
            'city': '',
            'airport': '',
            'cityCode': '',
            'country': '',
        }  # 定义列表tmp存储每一行的数据
        for ci, cell in enumerate(rows):  # 遍历一行每个单元格的数据
            if(ci == 0):
                if(cell.value and '\xa0-' in cell.value):
                    tmp['city'] = cell.value.split('\xa0-')[0]
                    tmp['airport'] = cell.value.split('\xa0-')[1]
                elif(cell.value and '-' in cell.value):
                    tmp['city'] = cell.value.split('-')[0]
                    tmp['airport'] = cell.value.split('-')[1]
                else:
                    tmp['city'] = cell.value
                    tmp['airport'] = cell.value
            elif(ci == 1):
                tmp['country'] = cell.value
            elif(ci == 2):
                tmp['cityCode'] = cell.value
            tmp['cityName'] = ''
        result.append(tmp)
    return result


cityInfoList = []
result = readCity1(r'D:\Github\LLTrip-Cruise\kitchen\3code.xlsx', 'city')
result2 = readCity2(
    r'D:\Github\LLTrip-Cruise\kitchen\airport&city&flight.xlsx', 'airport')

for k, v in enumerate(result2):
    for k1, v1 in enumerate(result):
        if(v["cityCode"] == v1["cityCode"]):
            v["cityName"] = v1["cityName"]

with open(r'D:\Github\LLTrip-Cruise\kitchen\cityInfo2.json', mode='w', encoding='utf-8') as jf:
    json.dump(result2, jf, indent=2, sort_keys=True, ensure_ascii=False)
